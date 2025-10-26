import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import sqlite3
import openpyxl
import random
import os
import time
import json 
from pathlib import Path
from datetime import datetime 
import hashlib # --- THÊM: Thư viện để băm ---

class LuckyDrawApp:
    def __init__(self, root):
        self.root = root
        
        # --- CẤU HÌNH ĐA NGÔN NGỮ (SỬ DỤNG JSON) ---
        self.config_data = {} # Sẽ chứa dữ liệu JSON
        self.config_file = Path('languages.json') # Đổi tên file config
        self.load_config() # Tải file config hoặc tạo file mặc định
        
        # Định nghĩa các ngôn ngữ hỗ trợ
        self.languages = {'en': 'English', 'vi': 'Tiếng Việt', 'zh-CN': '简体中文', 'zh-TW': '繁體中文'}
        # Map ngược từ tên hiển thị về mã code
        self.lang_map_display_to_code = {v: k for k, v in self.languages.items()}
        # Map từ mã code sang tên hiển thị
        self.lang_map_code_to_display = {k: v for k, v in self.languages.items()}

        # Biến lưu trữ mã ngôn ngữ hiện tại (vd: 'vi')
        self.current_lang_code = tk.StringVar(value='vi') 
        # Biến lưu trữ tên ngôn ngữ đang hiển thị (vd: 'Tiếng Việt')
        self.current_lang_display = tk.StringVar()
        self.current_lang_display.set(self.lang_map_code_to_display[self.current_lang_code.get()])
        # --- KẾT THÚC CẤU HÌNH ĐA NGÔN NGỮ ---

        self.root.title(self.lang("app_title"))
        self.root.geometry("800x650")

        # Cấu hình style cho ttk
        style = ttk.Style()
        style.theme_use('clam') 
        
        # --- Cấu hình chung cho TButton ---
        style.configure("TButton", padding=6, relief="raised", borderwidth=2, font=('Arial', 10))
        style.map("TButton", 
            background=[('active', 'grey'), ('!disabled', 'lightgrey')], 
            foreground=[('!disabled', 'black')] 
        )
        
        # --- Style riêng cho nút Lựa chọn (Draw) ---
        style.configure("Red.TButton", background="red", foreground="white", font=('Arial', 10, 'bold'))
        style.map("Red.TButton", 
            background=[('active', 'darkred'), ('!disabled', 'red')], 
            foreground=[('active', 'white'), ('!disabled', 'white')] 
        )

        style.configure("TLabel", padding=5, font=('Arial', 10))
        style.configure("TEntry", padding=5, font=('Arial', 10))
        style.configure("Treeview.Heading", font=('Arial', 10, 'bold'))

        # Khởi tạo cơ sở dữ liệu
        # --- THAY ĐỔI: Đổi tên DB cho mục đích chung ---
        self.db_name = 'random_selector.db'
        # --- KẾT THÚC THAY ĐỔI ---
        self.conn = None
        self.cursor = None
        self.setup_database()

        # Dữ liệu nội bộ
        self.available_employees = []
        self.winners = [] # Giữ tên biến nội bộ (winners) để tránh lỗi, nhưng UI sẽ hiển thị là "Selected"

        # --- Giao diện ---

        # Frame chính
        main_frame = ttk.Frame(root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Frame cho các nút điều khiển trên cùng
        top_frame = ttk.Frame(main_frame)
        top_frame.pack(fill=tk.X, pady=5)

        self.btn_load = ttk.Button(top_frame, text=self.lang("btn_load_excel"), command=self.load_excel)
        self.btn_load.pack(side=tk.LEFT, padx=5)

        self.btn_shuffle = ttk.Button(top_frame, text=self.lang("btn_shuffle"), command=self.shuffle_available_list)
        self.btn_shuffle.pack(side=tk.LEFT, padx=5)

        self.btn_reset = ttk.Button(top_frame, text=self.lang("btn_reset"), command=self.reset_draw)
        self.btn_reset.pack(side=tk.LEFT, padx=5)

        self.btn_clear_db = ttk.Button(top_frame, text=self.lang("btn_clear_db"), command=self.clear_all_data)
        self.btn_clear_db.pack(side=tk.LEFT, padx=5)

        # Bộ chọn ngôn ngữ
        self.lang_combo = ttk.Combobox(
            top_frame, 
            textvariable=self.current_lang_display,
            values=list(self.languages.values()),
            state="readonly",
            width=12
        )
        self.lang_combo.pack(side=tk.RIGHT, padx=5)
        self.lang_combo.bind("<<ComboboxSelected>>", self.on_language_change)
        
        self.label_language = ttk.Label(top_frame, text=self.lang("label_language"))
        self.label_language.pack(side=tk.RIGHT, padx=(10, 2))

        # Frame cho hai danh sách (Chờ và Đã chọn)
        list_frame = ttk.Frame(main_frame)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        # Pane bên trái: Danh sách chờ
        left_pane = ttk.Frame(list_frame, padding=5)
        left_pane.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)

        # --- THAY ĐỔI: Nhãn này sẽ được cập nhật với số lượng ---
        self.label_available = ttk.Label(left_pane, text=self.lang("label_waiting_list"), font=('Arial', 12, 'bold'))
        self.label_available.pack(pady=5)
        
        # Frame và Scrollbar cho cây Chờ
        scroll_frame_available = ttk.Frame(left_pane)
        scroll_frame_available.pack(fill=tk.BOTH, expand=True)
        
        self.scroll_available = ttk.Scrollbar(scroll_frame_available, orient=tk.VERTICAL)
        
        # --- THAY ĐỔI: Thêm cột STT ---
        cols_available = (self.lang('tree_col_stt'), self.lang('tree_col_id'), self.lang('tree_col_name'))
        
        self.tree_available = ttk.Treeview(scroll_frame_available, columns=cols_available, show='headings', yscrollcommand=self.scroll_available.set)
        
        self.scroll_available.config(command=self.tree_available.yview)
        self.scroll_available.pack(side=tk.RIGHT, fill=tk.Y)
        
        for col in cols_available:
            self.tree_available.heading(col, text=col)
        # --- THAY ĐỔI: Cấu hình độ rộng cột STT ---
        self.tree_available.column(cols_available[0], width=50, minwidth=40, stretch=tk.NO) 
        self.tree_available.column(cols_available[1], width=120)
        self.tree_available.column(cols_available[2], width=180)
        # --- KẾT THÚC THAY ĐỔI ---
        
        self.tree_available.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.btn_export_available = ttk.Button(left_pane, text=self.lang("btn_export_waiting"), command=self.export_available_list)
        self.btn_export_available.pack(pady=10)

        # Pane bên phải: Danh sách đã chọn
        right_pane = ttk.Frame(list_frame, padding=5)
        right_pane.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=5)

        # --- THAY ĐỔI: Nhãn này sẽ được cập nhật với số lượng ---
        self.label_winners = ttk.Label(right_pane, text=self.lang("label_winners_list"), font=('Arial', 12, 'bold'))
        self.label_winners.pack(pady=5)

        # Frame và Scrollbar cho cây Đã chọn
        scroll_frame_winners = ttk.Frame(right_pane)
        scroll_frame_winners.pack(fill=tk.BOTH, expand=True)
        
        self.scroll_winners = ttk.Scrollbar(scroll_frame_winners, orient=tk.VERTICAL)

        # --- THAY ĐỔI: Thêm cột STT vào DS Đã Chọn ---
        cols_winners = (self.lang('tree_col_stt'), self.lang('tree_col_id'), self.lang('tree_col_name'), self.lang('tree_col_prize'))
        
        self.tree_winners = ttk.Treeview(scroll_frame_winners, columns=cols_winners, show='headings', yscrollcommand=self.scroll_winners.set)
        
        self.scroll_winners.config(command=self.tree_winners.yview)
        self.scroll_winners.pack(side=tk.RIGHT, fill=tk.Y)
        
        for col in cols_winners:
            self.tree_winners.heading(col, text=col)
        # --- THAY ĐỔI: Cấu hình độ rộng 4 cột (thêm STT) ---
        self.tree_winners.column(cols_winners[0], width=50, minwidth=40, stretch=tk.NO) # Cột STT
        self.tree_winners.column(cols_winners[1], width=120)
        self.tree_winners.column(cols_winners[2], width=120)
        self.tree_winners.column(cols_winners[3], width=110)
        # --- KẾT THÚC THAY ĐỔI ---
            
        self.tree_winners.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.btn_export_winners = ttk.Button(right_pane, text=self.lang("btn_export_winners"), command=self.export_winners_list)
        self.btn_export_winners.pack(pady=10)

        # Frame cho hành động (dưới cùng)
        draw_frame = ttk.Frame(main_frame, padding=10)
        draw_frame.pack(fill=tk.X)

        # Frame con cho các nút và entry
        input_frame = ttk.Frame(draw_frame)
        input_frame.pack(fill=tk.X)

        self.label_draw_qty = ttk.Label(input_frame, text=self.lang("label_draw_qty"))
        self.label_draw_qty.pack(side=tk.LEFT, padx=5)
        self.entry_num = ttk.Entry(input_frame, width=10)
        self.entry_num.pack(side=tk.LEFT, padx=5)

        # --- THAY ĐỔI: Nhãn 'Tên Giải Thưởng' đổi thành 'Tên Nhóm' ---
        self.label_prize_name = ttk.Label(input_frame, text=self.lang("label_prize_name"))
        self.label_prize_name.pack(side=tk.LEFT, padx=(10, 5))
        self.entry_prize_name = ttk.Entry(input_frame, width=15)
        self.entry_prize_name.pack(side=tk.LEFT, padx=5)
        # --- KẾT THÚC THAY ĐỔI ---

        self.btn_draw = ttk.Button(input_frame, text=self.lang("btn_draw"), command=self.draw_winners, style="Red.TButton")
        self.btn_draw.pack(side=tk.LEFT, padx=10)

        # Frame con cho progress bar
        progress_frame = ttk.Frame(draw_frame)
        progress_frame.pack(fill=tk.X, expand=True, pady=10)

        self.label_progress = ttk.Label(progress_frame, text=self.lang("label_drawing"))
        self.label_progress.pack(anchor='w')
        self.progress_bar = ttk.Progressbar(progress_frame, orient='horizontal', mode='determinate', length=100)
        self.progress_bar.pack(fill=tk.X, expand=True)
        self.progress_bar["value"] = 0

        # Tải dữ liệu từ DB khi khởi động
        self.load_data_from_db()
        # Cập nhật UI lần đầu
        self.update_ui_text()

    # --- CÁC HÀM XỬ LÝ ĐA NGÔN NGỮ (ĐÃ CẬP NHẬT CHO JSON) ---

    def get_default_config_dict(self):
        """
        Trả về nội dung config JSON mặc định (ĐÃ CẬP NHẬT TÊN GỌI)
        Hàm này cũng tự động thêm các key mới (STT, Count)
        """
        # --- THAY ĐỔI: Thêm key 'export_hash_label' ---
        return {
          "en": {
            "app_title": "Random Selector Software",
            "label_language": "Language:",
            "btn_load_excel": "1. Load Excel File",
            "btn_shuffle": "2. Shuffle (Waiting List)",
            "btn_reset": "Reset Selection List",
            "btn_draw": "3. Select",
            "label_waiting_list": "Waiting List",
            "label_winners_list": "Selected List",
            "btn_export_waiting": "Export Waiting...",
            "btn_export_winners": "Export Selected...",
            "label_draw_qty": "Enter quantity to select:",
            "label_drawing": "Selecting...",
            "tree_col_id": "Emp ID",
            "tree_col_name": "Employee Name",
            "success_title": "Success",
            "error_title": "Error",
            "warning_title": "Warning",
            "confirm_title": "Confirm Reset",
            "load_success_msg": "Loaded and updated {count} employees into the database.\nThe entire list has been reset.",
            "load_error_msg": "Could not read Excel file: {error}",
            "db_error_connect": "Could not connect or create table: {error}",
            "db_error_load": "Could not load data: {error}",
            "db_error_update_winner": "Could not update selection status: {error}",
            "db_error_reset": "Could not reset list: {error}",
            "empty_list_warn": "No one in the waiting list to shuffle.",
            "shuffle_success": "Shuffled the waiting list.",
            "invalid_number_error": "Please enter a valid number.",
            "zero_draw_error": "Selection quantity must be greater than 0.",
            "not_enough_error": "Only {remaining} people left in the waiting list, cannot select {requested}.",
            "draw_success_msg": "Successfully selected {count} items!",
            "reset_confirm_msg": "Are you sure you want to reset ALL employees to the waiting list?\n(This will clear the current selected list)",
            "reset_success_msg": "Entire list has been reset.",
            "empty_export_warn": "No data to export.",
            "export_success_msg": "Successfully exported file to:\n{filepath}",
            "export_error_msg": "Could not save Excel file: {error}",
            "export_sheet_name": "List",
            "export_header_stt": "No.",
            "export_header_id": "Employee ID",
            "export_header_name": "Employee Name",
            "export_prefix_waiting": "Waiting_List",
            "export_prefix_winners": "Selected_List",
            "file_dialog_title_excel": "Select Excel file",
            "file_dialog_title_save": "Save Excel file",
            "btn_clear_db": "Clear All Data",
            "clear_all_confirm_msg": "WARNING: This will permanently delete ALL employee data from the database. This action cannot be undone.\n\nAre you sure?",
            "clear_all_success_msg": "All employee data has been successfully deleted.",
            "db_error_clear": "Could not clear database: {error}",
            "label_prize_name": "Group Name:",
            "tree_col_prize": "Group",
            "export_header_prize": "Group",
            "prize_name_empty_warn": "Please enter a group name to proceed.",
            "tree_col_stt": "No.",
            "label_waiting_list_count": "Waiting List (Total: {count})",
            "label_winners_list_count": "Selected List (Total: {count})",
            "export_hash_label": "Data Integrity Check (SHA256):"
          },
          "vi": {
            "app_title": "Phần mềm Sắp xếp Ngẫu nhiên",
            "label_language": "Ngôn ngữ:",
            "btn_load_excel": "1. Tải File Excel",
            "btn_shuffle": "2. Sắp xếp ngẫu nhiên (DS Chờ)",
            "btn_reset": "Reset DS Lựa chọn",
            "btn_draw": "3. Lựa chọn",
            "label_waiting_list": "Danh sách Chờ",
            "label_winners_list": "Danh sách Đã chọn",
            "btn_export_waiting": "Xuất DS Chờ...",
            "btn_export_winners": "Xuất DS Đã chọn...",
            "label_draw_qty": "Nhập số lượng muốn chọn:",
            "label_drawing": "Đang lựa chọn...",
            "tree_col_id": "Mã NV",
            "tree_col_name": "Tên Nhân Viên",
            "success_title": "Thành công",
            "error_title": "Lỗi",
            "warning_title": "Cảnh báo",
            "confirm_title": "Xác nhận Reset",
            "load_success_msg": "Đã tải và cập nhật {count} nhân viên vào cơ sở dữ liệu.\nToàn bộ danh sách đã được reset.",
            "load_error_msg": "Không thể đọc file Excel: {error}",
            "db_error_connect": "Không thể kết nối hoặc tạo bảng: {error}",
            "db_error_load": "Không thể tải dữ liệu: {error}",
            "db_error_update_winner": "Không thể cập nhật trạng thái đã chọn: {error}",
            "db_error_reset": "Không thể reset danh sách: {error}",
            "empty_list_warn": "Không có ai trong danh sách chờ để sắp xếp.",
            "shuffle_success": "Đã xáo trộn danh sách chờ.",
            "invalid_number_error": "Vui lòng nhập một con số hợp lệ.",
            "zero_draw_error": "Số lượng chọn phải lớn hơn 0.",
            "not_enough_error": "Chỉ còn {remaining} người trong danh sách chờ, không thể chọn {requested} người.",
            "draw_success_msg": "Đã chọn thành công {count} mục!",
            "reset_confirm_msg": "Bạn có chắc muốn reset TẤT CẢ nhân viên về danh sách chờ?\n(Hành động này sẽ xóa hết danh sách đã chọn hiện tại)",
            "reset_success_msg": "Đã reset toàn bộ danh sách.",
            "empty_export_warn": "Không có dữ liệu để xuất.",
            "export_success_msg": "Đã xuất file thành công tại:\n{filepath}",
            "export_error_msg": "Không thể lưu file Excel: {error}",
            "export_sheet_name": "Danh Sách",
            "export_header_stt": "STT",
            "export_header_id": "Mã Nhân Viên",
            "export_header_name": "Tên Nhân Viên",
            "export_prefix_waiting": "DS_Cho",
            "export_prefix_winners": "DS_DaChon",
            "file_dialog_title_excel": "Chọn file Excel",
            "file_dialog_title_save": "Lưu file Excel",
            "btn_clear_db": "Xóa Sạch Dữ Liệu",
            "clear_all_confirm_msg": "CẢNH BÁO: Hành động này sẽ XÓA VĨNH VIỄN toàn bộ dữ liệu nhân viên khỏi cơ sở dữ liệu. Không thể hoàn tác.\n\nBạn có chắc chắn không?",
            "clear_all_success_msg": "Đã xóa sạch toàn bộ dữ liệu nhân viên.",
            "db_error_clear": "Không thể xóa sạch cơ sở dữ liệu: {error}",
            "label_prize_name": "Tên nhóm:",
            "tree_col_prize": "Nhóm",
            "export_header_prize": "Nhóm",
            "prize_name_empty_warn": "Vui lòng nhập tên nhóm để tiếp tục.",
            "tree_col_stt": "STT",
            "label_waiting_list_count": "Danh sách Chờ (Tổng: {count})",
            "label_winners_list_count": "Danh sách Đã chọn (Tổng: {count})",
            "export_hash_label": "Mã Kiểm tra Toàn vẹn (SHA256):"
          },
          "zh-CN": {
            "app_title": "随机选择软件",
            "label_language": "语言:",
            "btn_load_excel": "1. 加载Excel文件",
            "btn_shuffle": "2. 随机排序 (等待列表)",
            "btn_reset": "重置选择列表",
            "btn_draw": "3. 选择",
            "label_waiting_list": "等待列表",
            "label_winners_list": "已选名单",
            "btn_export_waiting": "导出等待列表...",
            "btn_export_winners": "导出已选名单...",
            "label_draw_qty": "输入选择数量:",
            "label_drawing": "正在选择...",
            "tree_col_id": "员工ID",
            "tree_col_name": "员工姓名",
            "success_title": "成功",
            "error_title": "错误",
            "warning_title": "警告",
            "confirm_title": "确认重置",
            "load_success_msg": "已加载并更新 {count} 名员工到数据库。\n整个列表已重置。",
            "load_error_msg": "无法读取Excel文件: {error}",
            "db_error_connect": "无法连接或创建表: {error}",
            "db_error_load": "无法加载数据: {error}",
            "db_error_update_winner": "无法更新选择状态: {error}",
            "db_error_reset": "无法重置列表: {error}",
            "empty_list_warn": "等待列表中无人可排序。",
            "shuffle_success": "已随机排序等待列表。",
            "invalid_number_error": "请输入一个有效数字。",
            "zero_draw_error": "选择数量必须大于0。",
            "not_enough_error": "等待列表中只剩下 {remaining} 人，无法选择 {requested} 人。",
            "draw_success_msg": "成功抽出 {count} 个项目！",
            "reset_confirm_msg": "您确定要将所有员工重置回等待列表吗？\n(这将清除当前的选择名单)",
            "reset_success_msg": "整个列表已重置。",
            "empty_export_warn": "没有数据可导出。",
            "export_success_msg": "文件已成功导出到:\n{filepath}",
            "export_error_msg": "无法保存Excel文件: {error}",
            "export_sheet_name": "列表",
            "export_header_stt": "序号",
            "export_header_id": "员工ID",
            "export_header_name": "员工姓名",
            "export_prefix_waiting": "等待列表",
            "export_prefix_winners": "已选名单",
            "file_dialog_title_excel": "选择Excel文件",
            "file_dialog_title_save": "保存Excel文件",
            "btn_clear_db": "清除所有数据",
            "clear_all_confirm_msg": "警告：这将从数据库中永久删除所有员工数据。此操作无法撤销。\n\n您确定吗？",
            "clear_all_success_msg": "已成功删除所有员工数据。",
            "db_error_clear": "无法清除数据库：{error}",
            "label_prize_name": "分组名称:",
            "tree_col_prize": "分组",
            "export_header_prize": "分组",
            "prize_name_empty_warn": "请输入分组名称以继续。",
            "tree_col_stt": "序号",
            "label_waiting_list_count": "等待列表 (总数: {count})",
            "label_winners_list_count": "已选名单 (总数: {count})",
            "export_hash_label": "数据完整性校验 (SHA256):"
          },
          "zh-TW": {
            "app_title": "隨機選擇軟體",
            "label_language": "語言:",
            "btn_load_excel": "1. 載入Excel檔案",
            "btn_shuffle": "2. 隨機排序 (等待列表)",
            "btn_reset": "重設選擇列表",
            "btn_draw": "3. 選擇",
            "label_waiting_list": "等待列表",
            "label_winners_list": "已選名單",
            "btn_export_waiting": "匯出等待列表...",
            "btn_export_winners": "匯出已選名單...",
            "label_draw_qty": "輸入選擇數量:",
            "label_drawing": "正在選擇...",
            "tree_col_id": "員工ID",
            "tree_col_name": "員工姓名",
            "success_title": "成功",
            "error_title": "錯誤",
            "warning_title": "警告",
            "confirm_title": "確認重設",
            "load_success_msg": "已載入並更新 {count} 名員工到資料庫。\n整個列表已重設。",
            "load_error_msg": "無法讀取Excel檔案: {error}",
            "db_error_connect": "無法連接或建立資料表: {error}",
            "db_error_load": "無法載入資料: {error}",
            "db_error_update_winner": "無法更新選擇狀態: {error}",
            "db_error_reset": "無法重設列表: {error}",
            "empty_list_warn": "等待列表中無人可排序。",
            "shuffle_success": "已隨機排序等待列表。",
            "invalid_number_error": "請輸入一個有效數字。",
            "zero_draw_error": "選擇數量必須大於0。",
            "not_enough_error": "等待列表中只剩下 {remaining} 人，無法選擇 {requested} 人。",
            "draw_success_msg": "成功抽出 {count} 個項目！",
            "reset_confirm_msg": "您確定要將所有員工重設回等待列表嗎？\n(這將清除目前的選擇名單)",
            "reset_success_msg": "整個列表已重設。",
            "empty_export_warn": "沒有資料可匯出。",
            "export_success_msg": "檔案已成功匯出到:\n{filepath}",
            "export_error_msg": "無法儲存Excel檔案: {error}",
            "export_sheet_name": "列表",
            "export_header_stt": "序號",
            "export_header_id": "員工ID",
            "export_header_name": "員工姓名",
            "export_prefix_waiting": "等待列表",
            "export_prefix_winners": "已選名單",
            "file_dialog_title_excel": "選擇Excel檔案",
            "file_dialog_title_save": "儲存Excel檔案",
            "btn_clear_db": "清除所有資料",
            "clear_all_confirm_msg": "警告：這將從資料庫中永久刪除所有員工資料。此操作無法撤銷。\n\n您確定嗎？",
            "clear_all_success_msg": "已成功刪除所有員工資料。",
            "db_error_clear": "無法清除資料庫：{error}",
            "label_prize_name": "分組名稱:",
            "tree_col_prize": "分組",
            "export_header_prize": "分組",
            "prize_name_empty_warn": "請輸入分組名稱以繼續。",
            "tree_col_stt": "序號",
            "label_waiting_list_count": "等待列表 (總數: {count})",
            "label_winners_list_count": "已選名單 (總數: {count})",
            "export_hash_label": "資料完整性校驗 (SHA256):"
          }
        }
        # --- KẾT THÚC THAY ĐỔI ---

    def load_config(self):
        """Tải file languages.json, hoặc tạo file mặc định nếu không tồn tại."""
        if not self.config_file.exists():
            try:
                with open(self.config_file, 'w', encoding='utf-8') as f:
                    # Ghi file JSON mặc định (đã cập nhật)
                    json.dump(self.get_default_config_dict(), f, indent=2, ensure_ascii=False)
            except IOError as e:
                messagebox.showerror("Lỗi Config", f"Không thể tạo file languages.json: {e}")
                self.root.quit()
        
        try:
            # Đọc file JSON
            with open(self.config_file, 'r', encoding='utf-8') as f:
                self.config_data = json.load(f)
        except json.JSONDecodeError as e:
            messagebox.showerror("Lỗi Config", f"File languages.json bị lỗi: {e}\n\Vui lòng sửa file hoặc xóa đi để tạo lại.")
            self.root.quit()
        except Exception as e:
             messagebox.showerror("Lỗi Config", f"Không thể đọc file languages.json: {e}")
             self.root.quit()


    def lang(self, key, **kwargs):
        """Lấy chuỗi dịch từ config JSON. Hỗ trợ format chuỗi."""
        lang_code = self.current_lang_code.get()
        try:
            lang_dict = self.config_data.get(lang_code, {})
            text = lang_dict.get(key, f"_{key}_")
            
            # Fallback về 'en' nếu không tìm thấy key
            if text == f"_{key}_" and lang_code != 'en':
                lang_dict = self.config_data.get('en', {})
                text = lang_dict.get(key, f"_{key}_")

            # Tự động thêm key mới vào config_data nếu thiếu (từ default)
            # Điều này giúp khi cập nhật phần mềm, file json cũ không bị lỗi
            if text == f"_{key}_":
                default_dict_all_lang = self.get_default_config_dict()
                default_lang_dict = default_dict_all_lang.get(lang_code, default_dict_all_lang.get('en', {}))
                text = default_lang_dict.get(key, f"_{key}_")

            if kwargs:
                return text.format(**kwargs)
            return text
        except Exception:
            return f"_{key}_" 

    def on_language_change(self, *args):
        """Sự kiện khi chọn ngôn ngữ mới từ combobox."""
        display_name = self.current_lang_display.get()
        code = self.lang_map_display_to_code[display_name]
        self.current_lang_code.set(code)
        self.update_ui_text()

    def update_ui_text(self):
        """Cập nhật toàn bộ văn bản trên giao diện sang ngôn ngữ đã chọn."""
        
        self.root.title(self.lang("app_title"))
        
        self.btn_load.config(text=self.lang("btn_load_excel"))
        self.btn_shuffle.config(text=self.lang("btn_shuffle"))
        self.btn_reset.config(text=self.lang("btn_reset"))
        self.btn_export_available.config(text=self.lang("btn_export_waiting"))
        self.btn_export_winners.config(text=self.lang("btn_export_winners"))
        self.btn_draw.config(text=self.lang("btn_draw")) 
        
        self.label_language.config(text=self.lang("label_language"))
        
        # --- THAY ĐỔI: Cập nhật nhãn với số lượng ---
        self.label_available.config(text=self.lang("label_waiting_list_count", count=len(self.available_employees)))
        self.label_winners.config(text=self.lang("label_winners_list_count", count=len(self.winners)))
        # --- KẾT THÚC THAY ĐỔI ---
        
        self.label_draw_qty.config(text=self.lang("label_draw_qty"))
        self.label_progress.config(text=self.lang("label_drawing"))
        self.btn_clear_db.config(text=self.lang("btn_clear_db"))
        self.label_prize_name.config(text=self.lang("label_prize_name"))

        # --- THAY ĐỔI: Cập nhật cột Treeview (Đã tách biệt 2 cây) ---
        # Cây danh sách chờ (3 cột)
        cols_available = (self.lang('tree_col_stt'), self.lang('tree_col_id'), self.lang('tree_col_name'))
        self.tree_available.config(columns=cols_available)
        self.tree_available.heading(cols_available[0], text=cols_available[0])
        self.tree_available.heading(cols_available[1], text=cols_available[1])
        self.tree_available.heading(cols_available[2], text=cols_available[2])
        self.tree_available.column(cols_available[0], width=50, minwidth=40, stretch=tk.NO) 
        self.tree_available.column(cols_available[1], width=120)
        self.tree_available.column(cols_available[2], width=180)

        # --- THAY ĐỔI: Cây danh sách đã chọn (4 cột) ---
        cols_winners = (self.lang('tree_col_stt'), self.lang('tree_col_id'), self.lang('tree_col_name'), self.lang('tree_col_prize'))
        self.tree_winners.config(columns=cols_winners)
        self.tree_winners.heading(cols_winners[0], text=cols_winners[0])
        self.tree_winners.heading(cols_winners[1], text=cols_winners[1])
        self.tree_winners.heading(cols_winners[2], text=cols_winners[2])
        self.tree_winners.heading(cols_winners[3], text=cols_winners[3])
        self.tree_winners.column(cols_winners[0], width=50, minwidth=40, stretch=tk.NO) # Cột STT
        self.tree_winners.column(cols_winners[1], width=120)
        self.tree_winners.column(cols_winners[2], width=120)
        self.tree_winners.column(cols_winners[3], width=110)
        # --- KẾT THÚC THAY ĐỔI ---

        # --- THAY ĐỔI: Tải lại dữ liệu (có STT cho cả 2 DS) ---
        self.populate_tree(self.tree_available, self.available_employees, add_stt=True)
        # --- THAY ĐỔI: Thêm cờ (flag) reset_stt_on_group_change=True ---
        self.populate_tree(self.tree_winners, self.winners, add_stt=True, reset_stt_on_group_change=True)
        # --- KẾT THÚC THAY ĐỔI ---

    # --- KẾT THÚC CÁC HÀM ĐA NGÔN NGỮ ---

    def setup_database(self):
        """Khởi tạo kết nối DB và tạo/cập nhật bảng nếu chưa tồn tại"""
        try:
            self.conn = sqlite3.connect(self.db_name)
            self.cursor = self.conn.cursor()
            # --- THAY ĐỔI: Đổi tên cột cho mục đích chung ---
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS employees (
                    id TEXT PRIMARY KEY,
                    name TEXT,
                    is_selected INTEGER DEFAULT 0, 
                    selection_order INTEGER,
                    group_name TEXT
                )
            ''')
            
            # Logic nâng cấp DB cũ (nếu có)
            try:
                # Đổi tên cột cũ (nếu tồn tại)
                self.cursor.execute("ALTER TABLE employees RENAME COLUMN has_won TO is_selected")
                self.cursor.execute("ALTER TABLE employees RENAME COLUMN draw_order TO selection_order")
                self.cursor.execute("ALTER TABLE employees RENAME COLUMN prize_name TO group_name")
            except sqlite3.OperationalError as e:
                # Bỏ qua lỗi nếu cột không tồn tại hoặc đã được đổi tên
                # print(f"DB upgrade note: {e}")
                pass 
            # --- KẾT THÚC THAY ĐỔI ---

            self.conn.commit()
        except sqlite3.Error as e:
            messagebox.showerror(self.lang("error_title"), self.lang("db_error_connect", error=e))
            self.root.quit()

    def load_excel(self):
        """Mở dialog chọn file, đọc file Excel và nạp vào DB"""
        filepath = filedialog.askopenfilename(
            title=self.lang("file_dialog_title_excel"), 
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if not filepath:
            return

        try:
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active
            
            count = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row and row[0] is not None and row[1] is not None:
                    emp_id = str(row[0])
                    emp_name = str(row[1])
                    
                    # --- THAY ĐỔI: Cập nhật tên cột khi INSERT ---
                    self.cursor.execute(
                        "INSERT OR REPLACE INTO employees (id, name, is_selected, selection_order, group_name) VALUES (?, ?, 0, NULL, NULL)",
                        (emp_id, emp_name)
                    )
                    # --- KẾT THÚC THAY ĐỔI ---
                    count += 1
            
            self.conn.commit()
            messagebox.showinfo(self.lang("success_title"), self.lang("load_success_msg", count=count))
            self.load_data_from_db()

        except Exception as e:
            messagebox.showerror(self.lang("error_title"), self.lang("load_error_msg", error=e))

    def load_data_from_db(self):
        """Tải dữ liệu từ DB vào hai danh sách (available và winners) và cập nhật Treeview"""
        try:
            # --- THAY ĐỔI: Cập nhật tên cột khi SELECT ---
            self.cursor.execute("SELECT id, name FROM employees WHERE is_selected = 0")
            self.available_employees = self.cursor.fetchall()
            
            self.cursor.execute("SELECT id, name, group_name FROM employees WHERE is_selected = 1 ORDER BY selection_order")
            self.winners = self.cursor.fetchall()
            # --- KẾT THÚC THAY ĐỔI ---

            # --- THAY ĐỔI: Cập nhật tiêu đề (label) với số lượng ---
            self.label_available.config(text=self.lang("label_waiting_list_count", count=len(self.available_employees)))
            self.label_winners.config(text=self.lang("label_winners_list_count", count=len(self.winners)))
            # --- KẾT THÚC THAY ĐỔI ---

            # --- THAY ĐỔI: Gọi populate_tree với add_stt=True cho cả 2 DS ---
            self.populate_tree(self.tree_available, self.available_employees, add_stt=True)
            # --- THAY ĐỔI: Thêm cờ (flag) reset_stt_on_group_change=True ---
            self.populate_tree(self.tree_winners, self.winners, add_stt=True, reset_stt_on_group_change=True)
            # --- KẾT THÚC THAY ĐỔI ---
            
        except sqlite3.Error as e:
            messagebox.showerror(self.lang("error_title"), self.lang("db_error_load", error=e))

    # --- THAY ĐỔI: Thêm tham số reset_stt_on_group_change ---
    def populate_tree(self, tree, data, add_stt=False, reset_stt_on_group_change=False):
        """Hiển thị dữ liệu lên một Treeview, có tùy chọn thêm STT"""
        for i in tree.get_children():
            tree.delete(i)
        
        if not add_stt:
            for item in data:
                tree.insert('', 'end', values=item)
            return

        # Logic for add_stt=True
        if not reset_stt_on_group_change:
            # Logic cũ: STT liên tục (dùng cho DS Chờ)
            for index, item in enumerate(data, 1):
                tree.insert('', 'end', values=(index,) + item)
        else:
            # Logic mới: Reset STT khi 'group_name' thay đổi (dùng cho DS Đã chọn)
            # 'data' cho winners là: (id, name, group_name)
            # Cột 'group_name' là item[2]
            current_group = None
            group_stt = 1
            for item in data:
                # item[2] là group_name
                if item[2] != current_group:
                    current_group = item[2]
                    group_stt = 1 # Reset STT
                
                tree.insert('', 'end', values=(group_stt,) + item)
                group_stt += 1
    # --- KẾT THÚC THAY ĐỔI ---


    def shuffle_available_list(self):
        """Xáo trộn danh sách chờ và cập nhật Treeview"""
        if not self.available_employees:
            messagebox.showwarning(self.lang("warning_title"), self.lang("empty_list_warn"))
            return
            
        random.shuffle(self.available_employees)
        # --- THAY ĐỔI: Gọi populate_tree với add_stt=True ---
        self.populate_tree(self.tree_available, self.available_employees, add_stt=True)
        # --- KẾT THÚC THAY ĐỔI ---
        messagebox.showinfo(self.lang("success_title"), self.lang("shuffle_success"))

    def draw_winners(self):
        """Chọn ngẫu nhiên N người từ danh sách chờ (với progress bar)"""
        try:
            num_to_draw = int(self.entry_num.get())
        except ValueError:
            messagebox.showerror(self.lang("error_title"), self.lang("invalid_number_error"))
            return

        # --- THAY ĐỔI: Kiểm tra tên nhóm (thay vì tên giải) ---
        prize_name = self.entry_prize_name.get().strip()
        if not prize_name:
            messagebox.showwarning(self.lang("warning_title"), self.lang("prize_name_empty_warn"))
            return
        # --- KẾT THÚC THAY ĐỔI ---

        if num_to_draw <= 0:
            messagebox.showerror(self.lang("error_title"), self.lang("zero_draw_error"))
            return
            
        if num_to_draw > len(self.available_employees):
            messagebox.showerror(self.lang("error_title"), self.lang("not_enough_error", remaining=len(self.available_employees), requested=num_to_draw))
            return

        self.btn_draw.config(state=tk.DISABLED)
        
        try:
            # ... (Phần Progress bar giữ nguyên) ...
            self.progress_bar["value"] = 0
            max_steps = 100
            simulation_time_seconds = 2.5
            
            for i in range(max_steps + 1):
                self.progress_bar["value"] = i
                self.root.update_idletasks()
                time.sleep(simulation_time_seconds / max_steps)
            
            new_winners = random.sample(self.available_employees, num_to_draw)
            
            try:
                # --- THAY ĐỔI: Cập nhật DB với tên cột đã đổi ---
                
                # 1. Lấy thứ tự chọn cao nhất hiện tại
                self.cursor.execute("SELECT MAX(selection_order) FROM employees WHERE selection_order IS NOT NULL")
                max_order_result = self.cursor.fetchone()
                current_max_order = max_order_result[0] if max_order_result[0] is not None else 0
                
                for i, winner in enumerate(new_winners, 1):
                    emp_id = winner[0]
                    new_draw_order = current_max_order + i
                    self.cursor.execute(
                        "UPDATE employees SET is_selected = 1, selection_order = ?, group_name = ? WHERE id = ?", 
                        (new_draw_order, prize_name, emp_id)
                    )
                # --- KẾT THÚC THAY ĐỔI ---
                
                self.conn.commit()
                messagebox.showinfo(self.lang("success_title"), self.lang("draw_success_msg", count=len(new_winners)))
                self.load_data_from_db() 

            except sqlite3.Error as e:
                # --- THAY ĐỔI: Cập nhật key báo lỗi ---
                messagebox.showerror(self.lang("error_title"), self.lang("db_error_update_winner", error=e))
                # --- KẾT THÚC THAY ĐỔI ---
        
        finally:
            self.progress_bar["value"] = 0
            self.root.update_idletasks()
            self.btn_draw.config(state=tk.NORMAL)

    def reset_draw(self):
        """Chuyển tất cả mọi người về trạng thái chưa chọn"""
        if not messagebox.askyesno(self.lang("confirm_title"), self.lang("reset_confirm_msg")):
            return
            
        try:
            # --- THAY ĐỔI: Reset các cột đã đổi tên ---
            self.cursor.execute("UPDATE employees SET is_selected = 0, selection_order = NULL, group_name = NULL")
            # --- KẾT THÚC THAY ĐỔI ---
            self.conn.commit()
            self.load_data_from_db()
            messagebox.showinfo(self.lang("success_title"), self.lang("reset_success_msg"))
        except sqlite3.Error as e:
            messagebox.showerror(self.lang("error_title"), self.lang("db_error_reset", error=e))

    def clear_all_data(self):
        """Xóa TOÀN BỘ dữ liệu khỏi bảng employees trong DB"""
        if not messagebox.askyesno(self.lang("warning_title"), self.lang("clear_all_confirm_msg")):
            return
            
        try:
            self.cursor.execute("DELETE FROM employees")
            self.conn.commit()
            self.load_data_from_db()
            messagebox.showinfo(self.lang("success_title"), self.lang("clear_all_success_msg"))
        except sqlite3.Error as e:
            messagebox.showerror(self.lang("error_title"), self.lang("db_error_clear", error=e))

    def export_list(self, data, filename_prefix_key):
        """Hàm chung để xuất một danh sách (data) ra file Excel, có thêm STT và Mã hash"""
        if not data:
            messagebox.showwarning(self.lang("warning_title"), self.lang("empty_export_warn"))
            return

        filename_prefix = self.lang(filename_prefix_key)
        current_time = datetime.now().strftime("%Y%m%d%H%M%S")

        filepath = filedialog.asksaveasfilename(
            title=self.lang("file_dialog_title_save"), 
            defaultextension=".xlsx",
            initialfile=f"{filename_prefix}_{current_time}.xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        
        if not filepath:
            return 

        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = self.lang("export_sheet_name") 
            
            # --- THAY ĐỔI: Logic tiêu đề giữ nguyên (đã cập nhật ở lần trước) ---
            if data and len(data[0]) == 3: # Đây là danh sách đã chọn (ID, Name, Group)
                sheet.append([
                    self.lang("export_header_stt"), 
                    self.lang("export_header_id"), 
                    self.lang("export_header_name"),
                    self.lang("export_header_prize") 
                ])
            else: # Đây là danh sách chờ (ID, Name)
                sheet.append([
                    self.lang("export_header_stt"), 
                    self.lang("export_header_id"), 
                    self.lang("export_header_name")
                ])
            # --- KẾT THÚC THAY ĐỔI ---
            
            # --- THAY ĐỔI: Logic STT động khi xuất file (để reset STT theo nhóm) ---
            data_string_for_hash = ""
            
            if filename_prefix_key == "export_prefix_winners":
                # Logic mới: Reset STT khi 'group_name' thay đổi (dùng cho DS Đã chọn)
                # 'data' (row) là: (id, name, group_name)
                # Cột 'group_name' là row[2]
                current_group = None
                group_stt = 1
                for row in data:
                    if row[2] != current_group:
                        current_group = row[2]
                        group_stt = 1 # Reset STT
                    
                    row_to_write = (group_stt,) + row # (STT, ID, Name, Group)
                    sheet.append(row_to_write)
                    data_string_for_hash += "".join(map(str, row_to_write))
                    group_stt += 1
            else:
                # Logic cũ: STT liên tục (dùng cho DS Chờ)
                for index, row in enumerate(data, 1):
                    # row là (ID, Name)
                    row_to_write = (index,) + row # (STT, ID, Name)
                    sheet.append(row_to_write) 
                    data_string_for_hash += "".join(map(str, row_to_write))
            
            # Sau khi ghi hết dữ liệu, tính hash
            hasher = hashlib.sha256()
            hasher.update(data_string_for_hash.encode('utf-8')) # Băm dưới dạng bytes
            hash_value = hasher.hexdigest()
            
            # Ghi hash vào cuối file
            sheet.append([]) # Thêm một dòng trống
            sheet.append([self.lang("export_hash_label"), hash_value])
            # --- KẾT THÚC THAY ĐỔI ---
                
            workbook.save(filepath)
            messagebox.showinfo(self.lang("success_title"), self.lang("export_success_msg", filepath=filepath))
        except Exception as e:
            messagebox.showerror(self.lang("error_title"), self.lang("export_error_msg", error=e))

    def export_available_list(self):
        """Gọi hàm xuất cho danh sách chờ (sử dụng key prefix)"""
        # Data là: (ID, Name)
        self.export_list(self.available_employees, "export_prefix_waiting")

    def export_winners_list(self):
        """Gọi hàm xuất cho danh sách đã chọn (sử dụng key prefix)"""
        # Data là: (ID, Name, Group)
        self.export_list(self.winners, "export_prefix_winners")

    def on_closing(self):
        """Đóng kết nối DB khi thoát ứng dụng"""
        if self.conn:
            self.conn.close()
        self.root.destroy()

if __name__ == "__main__":
    root = tk.Tk()
    app = LuckyDrawApp(root)
    root.protocol("WM_DELETE_WINDOW", app.on_closing) 
    root.mainloop()




